#!/usr/bin/env python
"""
T51（パック組み合わせ情報）をContractPackテーブルにインポート
"""
import os
import sys
import uuid
from pathlib import Path

# Django setup
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

import django
django.setup()

from django.db import transaction
from openpyxl import load_workbook
from apps.contracts.models import Product, ContractPack

# テナントID（デフォルト）
DEFAULT_TENANT_ID = uuid.UUID('00000000-0000-0000-0000-000000000001')


def import_t51_to_contractpack(xlsx_path, tenant_id=DEFAULT_TENANT_ID):
    """T51 ExcelをContractPackにインポート"""

    print(f"Importing T51 data to ContractPack from: {xlsx_path}")
    print(f"Tenant ID: {tenant_id}")

    stats = {
        'total_rows': 0,
        'packs_created': 0,
        'packs_updated': 0,
        'skipped': 0,
        'errors': []
    }

    wb = load_workbook(xlsx_path)
    ws = wb.active

    print(f"Total rows to process: {ws.max_row - 1}")

    # 商品コードをキャッシュ
    product_cache = {}
    for product in Product.objects.filter(tenant_id=tenant_id):
        product_cache[product.product_code] = product

    print(f"Products in cache: {len(product_cache)}")

    with transaction.atomic():
        for row in ws.iter_rows(min_row=2, values_only=True):
            stats['total_rows'] += 1

            if stats['total_rows'] % 1000 == 0:
                print(f"  Processing row {stats['total_rows']}...")

            try:
                pack_id = row[0]  # パック契約ID
                base_id_1 = row[1] if len(row) > 1 else None
                base_id_2 = row[2] if len(row) > 2 else None
                base_id_3 = row[3] if len(row) > 3 else None
                base_id_4 = row[4] if len(row) > 4 else None

                if not pack_id:
                    stats['skipped'] += 1
                    continue

                # パック商品を取得
                pack_product = product_cache.get(pack_id)
                if not pack_product:
                    stats['errors'].append({
                        'row': stats['total_rows'] + 1,
                        'pack_id': pack_id,
                        'error': 'Pack product not found'
                    })
                    stats['skipped'] += 1
                    continue

                # 基本商品を取得
                base_product_1 = product_cache.get(base_id_1) if base_id_1 else None
                base_product_2 = product_cache.get(base_id_2) if base_id_2 else None
                base_product_3 = product_cache.get(base_id_3) if base_id_3 else None
                base_product_4 = product_cache.get(base_id_4) if base_id_4 else None

                # 少なくとも1つの基本商品が必要
                if not base_product_1:
                    stats['skipped'] += 1
                    continue

                # ContractPackを作成または更新
                contract_pack, created = ContractPack.objects.update_or_create(
                    tenant_id=tenant_id,
                    pack_product=pack_product,
                    defaults={
                        'base_product_1': base_product_1,
                        'base_product_2': base_product_2,
                        'base_product_3': base_product_3,
                        'base_product_4': base_product_4,
                    }
                )

                if created:
                    stats['packs_created'] += 1
                else:
                    stats['packs_updated'] += 1

            except Exception as e:
                stats['errors'].append({
                    'row': stats['total_rows'] + 1,
                    'error': str(e)
                })

    # 結果表示
    print("\n" + "=" * 50)
    print("Import Complete!")
    print("=" * 50)
    print(f"Total rows processed: {stats['total_rows']}")
    print(f"ContractPacks created: {stats['packs_created']}")
    print(f"ContractPacks updated: {stats['packs_updated']}")
    print(f"Skipped: {stats['skipped']}")
    print(f"Errors: {len(stats['errors'])}")

    if stats['errors']:
        print("\nFirst 10 errors:")
        for err in stats['errors'][:10]:
            print(f"  - {err}")

    return stats


if __name__ == '__main__':
    xlsx_path = '/Users/hirosesuzu/Desktop/アンシステム/Claude-Code-Communication/instructions/おざ/T51_パック組み合わせ情報.xlsx'

    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]

    import_t51_to_contractpack(xlsx_path)
