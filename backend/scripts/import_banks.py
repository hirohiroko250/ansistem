"""
銀行マスタデータインポートスクリプト
東海圏の金融機関と支店をインポート
使用方法: docker compose run --rm backend python scripts/import_banks.py <xlsx_path>
"""
import sys
import os
import django
import openpyxl

# Django設定
sys.path.insert(0, '/app')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.development')
django.setup()

from apps.schools.models import Bank, BankBranch, BankType
from apps.tenants.models import Tenant


# 東海圏で使われる可能性が高い銀行コード
IMPORTANT_BANK_CODES = {
    # メガバンク
    '0001': 'みずほ銀行',
    '0005': '三菱UFJ銀行',
    '0009': '三井住友銀行',
    '0010': 'りそな銀行',
    # ゆうちょ
    '9900': 'ゆうちょ銀行',
    # 東海の地方銀行
    '0149': '静岡銀行',
    '0150': '大垣共立銀行',
    '0151': '十六銀行',
    '0152': '三重銀行',
    '0153': '百五銀行',
    '0538': '静岡中央銀行',
    '0542': '愛知銀行',
    '0543': '名古屋銀行',
    '0544': '中京銀行',
}

# 東海圏キーワード
TOKAI_KEYWORDS = [
    '愛知', '名古屋', '岐阜', '三重', '静岡', '東海',
    '豊橋', '岡崎', '豊田', '一宮', '半田', '瀬戸',
    'あいち', '碧海', '知多', '尾西', '蒲郡', '西尾',
    '刈谷', '安城', '豊川', '桑名', '四日市', '浜松', '沼津',
    '中日',  # 中日信用金庫
]


def get_aiueo_row(hiragana):
    """ひらがなからあいうえお行を取得"""
    if not hiragana:
        return ''
    first_char = hiragana[0]
    # 濁音・半濁音を清音に変換
    dakuon_map = {
        'が': 'か', 'ぎ': 'か', 'ぐ': 'か', 'げ': 'か', 'ご': 'か',
        'ざ': 'さ', 'じ': 'さ', 'ず': 'さ', 'ぜ': 'さ', 'ぞ': 'さ',
        'だ': 'た', 'ぢ': 'た', 'づ': 'た', 'で': 'た', 'ど': 'た',
        'ば': 'は', 'び': 'は', 'ぶ': 'は', 'べ': 'は', 'ぼ': 'は',
        'ぱ': 'は', 'ぴ': 'は', 'ぷ': 'は', 'ぺ': 'は', 'ぽ': 'は',
    }
    first_char = dakuon_map.get(first_char, first_char)

    if first_char in 'あいうえお':
        return 'あ'
    elif first_char in 'かきくけこ':
        return 'か'
    elif first_char in 'さしすせそ':
        return 'さ'
    elif first_char in 'たちつてと':
        return 'た'
    elif first_char in 'なにぬねの':
        return 'な'
    elif first_char in 'はひふへほ':
        return 'は'
    elif first_char in 'まみむめも':
        return 'ま'
    elif first_char in 'やゆよ':
        return 'や'
    elif first_char in 'らりるれろ':
        return 'ら'
    elif first_char in 'わをん':
        return 'わ'
    return ''


def katakana_to_hiragana(text):
    """カタカナをひらがなに変換"""
    if not text:
        return ''
    result = []
    for char in text:
        code = ord(char)
        if 0x30A1 <= code <= 0x30F6:  # カタカナ範囲
            result.append(chr(code - 0x60))  # ひらがなに変換
        else:
            result.append(char)
    return ''.join(result)


def is_tokai_bank(bank_name, bank_code):
    """東海圏の銀行かどうか判定"""
    bank_code_str = str(bank_code)
    # '旧'で始まるコードは旧銀行なのでスキップ
    if bank_code_str.startswith('旧'):
        return False
    bank_code = bank_code_str.zfill(4)
    if bank_code in IMPORTANT_BANK_CODES:
        return True
    for kw in TOKAI_KEYWORDS:
        if kw in bank_name:
            return True
    return False


def import_banks(xlsx_path):
    """銀行と支店をインポート"""

    # テナント取得
    tenant = Tenant.objects.filter(tenant_code='100000').first()
    if not tenant:
        print("エラー: テナントが見つかりません")
        sys.exit(1)
    tenant_id = tenant.id
    print(f"テナント: {tenant.tenant_name} ({tenant_id})")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb['全銀協コード (支店名)']

    # 対象銀行コードを収集
    target_bank_codes = set()
    bank_info = {}  # bank_code -> (bank_name, bank_name_kana)

    for row in ws.iter_rows(min_row=2):
        bank_name = str(row[1].value).strip() if row[1].value else ''
        bank_name_kana = str(row[3].value).strip() if row[3].value else ''
        bank_code = str(row[6].value).strip().zfill(4) if row[6].value else ''

        if not bank_code or not bank_name:
            continue

        if is_tokai_bank(bank_name, bank_code):
            target_bank_codes.add(bank_code)
            if bank_code not in bank_info:
                bank_info[bank_code] = (bank_name, bank_name_kana)

    print(f"対象銀行数: {len(target_bank_codes)}")

    # 銀行を作成
    banks_created = 0
    bank_map = {}  # bank_code -> Bank instance

    for bank_code in sorted(target_bank_codes):
        bank_name, bank_name_kana = bank_info.get(bank_code, ('', ''))
        if not bank_name:
            continue

        # ひらがな変換
        bank_name_hiragana = katakana_to_hiragana(bank_name_kana)

        existing = Bank.objects.filter(tenant_id=tenant_id, bank_code=bank_code).first()
        if existing:
            bank_map[bank_code] = existing
            continue

        # bank_codeから数字部分のみ抽出してsort_orderにする
        numeric_part = ''.join(c for c in bank_code if c.isdigit())
        sort_order_val = int(numeric_part) if numeric_part else 0

        bank = Bank.objects.create(
            tenant_id=tenant_id,
            bank_code=bank_code,
            bank_name=bank_name,
            bank_name_kana=bank_name_kana,
            bank_name_half_kana=bank_name_kana,
            bank_name_hiragana=bank_name_hiragana,
            aiueo_row=get_aiueo_row(bank_name_hiragana),
            sort_order=sort_order_val,
            is_active=True,
        )
        bank_map[bank_code] = bank
        banks_created += 1

    print(f"銀行作成: {banks_created} 件")

    # 支店を作成
    branches_created = 0
    branches_updated = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            bank_code = str(row[6].value).strip().zfill(4) if row[6].value else ''
            if bank_code not in target_bank_codes:
                continue

            bank = bank_map.get(bank_code)
            if not bank:
                continue

            branch_name = str(row[4].value).strip() if row[4].value else ''
            branch_name_kana = str(row[5].value).strip() if row[5].value else ''
            branch_code = str(row[7].value).strip().zfill(3) if row[7].value else ''

            if not branch_code or not branch_name:
                continue

            # ひらがな変換
            branch_name_hiragana = katakana_to_hiragana(branch_name_kana)

            existing = BankBranch.objects.filter(bank=bank, branch_code=branch_code).first()
            if existing:
                # 更新
                existing.branch_name = branch_name
                existing.branch_name_kana = branch_name_kana
                existing.branch_name_half_kana = branch_name_kana
                existing.branch_name_hiragana = branch_name_hiragana
                existing.aiueo_row = get_aiueo_row(branch_name_hiragana)
                existing.save()
                branches_updated += 1
            else:
                # branch_codeから数字部分のみ抽出
                branch_numeric = ''.join(c for c in branch_code if c.isdigit())
                branch_sort = int(branch_numeric) if branch_numeric else 0
                BankBranch.objects.create(
                    tenant_id=tenant_id,
                    bank=bank,
                    branch_code=branch_code,
                    branch_name=branch_name,
                    branch_name_kana=branch_name_kana,
                    branch_name_half_kana=branch_name_kana,
                    branch_name_hiragana=branch_name_hiragana,
                    aiueo_row=get_aiueo_row(branch_name_hiragana),
                    sort_order=branch_sort,
                    is_active=True,
                )
                branches_created += 1

            if (branches_created + branches_updated) % 500 == 0:
                print(f"  処理中... {branches_created + branches_updated} 件")

        except Exception as e:
            errors.append(f"行 {row_num}: {str(e)}")
            continue

    return banks_created, branches_created, branches_updated, errors


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("使用方法: python scripts/import_banks.py <xlsx_path>")
        sys.exit(1)

    xlsx_path = sys.argv[1]

    if not os.path.exists(xlsx_path):
        print(f"エラー: ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    print(f"インポート開始: {xlsx_path}")
    print("-" * 50)

    banks_created, branches_created, branches_updated, errors = import_banks(xlsx_path)

    print("-" * 50)
    print(f"完了!")
    print(f"  銀行新規: {banks_created} 件")
    print(f"  支店新規: {branches_created} 件")
    print(f"  支店更新: {branches_updated} 件")

    if errors:
        print(f"  エラー: {len(errors)} 件")
        for error in errors[:10]:
            print(f"    - {error}")
        if len(errors) > 10:
            print(f"    ... 他 {len(errors) - 10} 件")
