#!/usr/bin/env python
"""
T51（パック組み合わせ情報）Excelインポートスクリプト

パック契約と基本契約の組み合わせ情報をインポート
"""
import os
import sys
import uuid
from pathlib import Path

# Django setup
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')

import django
django.setup()

from django.db import transaction
from openpyxl import load_workbook
from apps.contracts.models import Product, ProductComposition

# テナントID（デフォルト）
DEFAULT_TENANT_ID = uuid.UUID('00000000-0000-0000-0000-000000000001')


def import_t51_excel(xlsx_path, tenant_id=DEFAULT_TENANT_ID):
    """T51 Excelをインポート"""

    print(f"Importing T51 data from: {xlsx_path}")
    print(f"Tenant ID: {tenant_id}")

    stats = {
        'total_rows': 0,
        'compositions_created': 0,
        'compositions_updated': 0,
        'skipped': 0,
        'errors': []
    }

    wb = load_workbook(xlsx_path)
    ws = wb.active

    print(f"Total rows to process: {ws.max_row - 1}")  # ヘッダー除く

    # 商品コードをキャッシュ
    product_cache = {}
    for product in Product.objects.filter(tenant_id=tenant_id):
        product_cache[product.product_code] = product

    print(f"Products in cache: {len(product_cache)}")

    with transaction.atomic():
        for row in ws.iter_rows(min_row=2, values_only=True):  # ヘッダースキップ
            stats['total_rows'] += 1

            if stats['total_rows'] % 1000 == 0:
                print(f"  Processing row {stats['total_rows']}...")

            try:
                pack_id = row[0]  # パック契約ID
                base_id_1 = row[1]  # 基本契約ID_1
                base_id_2 = row[2] if len(row) > 2 else None  # 基本契約ID_2
                base_id_3 = row[3] if len(row) > 3 else None  # 基本契約ID_3
                base_id_4 = row[4] if len(row) > 4 else None  # 基本契約ID_4

                if not pack_id:
                    stats['skipped'] += 1
                    continue

                # パック商品を取得
                pack_product = product_cache.get(pack_id)
                if not pack_product:
                    stats['errors'].append({
                        'row': stats['total_rows'] + 1,
                        'pack_id': pack_id,
                        'error': 'Pack product not found'
                    })
                    stats['skipped'] += 1
                    continue

                # 基本契約を取得してProductCompositionを作成
                base_ids = [base_id_1, base_id_2, base_id_3, base_id_4]
                sort_order = 0

                for base_id in base_ids:
                    if not base_id or base_id == '':
                        continue

                    base_product = product_cache.get(base_id)
                    if not base_product:
                        # 商品が見つからない場合はスキップ（エラーログは出さない）
                        continue

                    # パック商品と基本商品が同じ場合はスキップ
                    if pack_product.id == base_product.id:
                        continue

                    sort_order += 1

                    # ProductCompositionを作成または更新
                    composition, created = ProductComposition.objects.update_or_create(
                        tenant_id=tenant_id,
                        parent_product=pack_product,
                        child_product=base_product,
                        defaults={
                            'quantity': 1,
                            'sort_order': sort_order,
                            'is_active': True
                        }
                    )

                    if created:
                        stats['compositions_created'] += 1
                    else:
                        stats['compositions_updated'] += 1

            except Exception as e:
                stats['errors'].append({
                    'row': stats['total_rows'] + 1,
                    'error': str(e)
                })

    # 結果表示
    print("\n" + "=" * 50)
    print("Import Complete!")
    print("=" * 50)
    print(f"Total rows processed: {stats['total_rows']}")
    print(f"Compositions created: {stats['compositions_created']}")
    print(f"Compositions updated: {stats['compositions_updated']}")
    print(f"Skipped: {stats['skipped']}")
    print(f"Errors: {len(stats['errors'])}")

    if stats['errors']:
        print("\nFirst 10 errors:")
        for err in stats['errors'][:10]:
            print(f"  - Row {err.get('row', '?')}: {err.get('error', 'unknown')}")

    return stats


if __name__ == '__main__':
    xlsx_path = '/Users/hirosesuzu/Desktop/アンシステム/Claude-Code-Communication/instructions/おざ/T51_パック組み合わせ情報.xlsx'

    if len(sys.argv) > 1:
        xlsx_path = sys.argv[1]

    import_t51_excel(xlsx_path)
