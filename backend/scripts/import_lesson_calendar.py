"""
開講カレンダーデータインポートスクリプト (T13)
使用方法: docker compose run --rm backend python scripts/import_lesson_calendar.py <xlsx_path>
"""
import sys
import os
import django
import openpyxl
from datetime import datetime

# Django設定
sys.path.insert(0, '/app')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.development')
django.setup()

from apps.schools.models import LessonCalendar, CalendarMaster, Brand
from apps.tenants.models import Tenant


def parse_date(value):
    """日付パース"""
    if not value:
        return None
    if isinstance(value, datetime):
        return value.date()
    try:
        return datetime.strptime(str(value)[:10], '%Y-%m-%d').date()
    except:
        return None


def detect_lesson_type(display_label, ticket_type):
    """授業タイプを検出"""
    if not display_label or display_label == '休':
        return LessonCalendar.LessonType.CLOSED

    # チケット券種からタイプ判定
    if ticket_type:
        ticket_upper = str(ticket_type).upper()
        if ticket_upper == 'A':
            return LessonCalendar.LessonType.TYPE_A
        elif ticket_upper == 'B':
            return LessonCalendar.LessonType.TYPE_B
        elif ticket_upper == 'P':
            return LessonCalendar.LessonType.TYPE_P
        elif ticket_upper == 'Y':
            return LessonCalendar.LessonType.TYPE_Y

    # 表示ラベルからタイプ判定
    if display_label:
        label = str(display_label).upper()
        if 'A' in label:
            return LessonCalendar.LessonType.TYPE_A
        elif 'B' in label:
            return LessonCalendar.LessonType.TYPE_B

    return LessonCalendar.LessonType.CLOSED


def import_calendar_sheet(ws, tenant_id, brand_map, calendar_master_map):
    """カレンダーシートをインポート"""
    imported = 0
    updated = 0
    errors = []

    for row_num, row in enumerate(ws.iter_rows(min_row=5), start=5):
        try:
            calendar_id = str(row[0].value).strip() if row[0].value else ''
            if not calendar_id:
                continue

            lesson_date = parse_date(row[1].value)
            if not lesson_date:
                continue

            day_of_week = str(row[2].value).strip() if row[2].value else ''
            display_label = str(row[3].value).strip() if row[3].value else ''
            is_open_val = str(row[4].value).strip() if row[4].value else ''
            rejection_code = str(row[5].value).strip() if row[5].value else ''
            ticket_type = str(row[6].value).strip() if row[6].value else ''
            ticket_issue_count = row[7].value
            ticket_sequence = row[8].value
            valid_days = row[9].value
            notice = str(row[10].value).strip() if row[10].value else ''
            auto_send = str(row[11].value).strip() if row[11].value else ''
            rejection_reason = str(row[12].value).strip() if row[12].value else ''
            holiday_name = str(row[13].value).strip() if row[13].value else ''

            # 開講判定
            is_open = is_open_val == 'Y' or (display_label and display_label != '休')

            # レッスンタイプ検出
            lesson_type = detect_lesson_type(display_label, ticket_type)

            # カレンダーマスター取得または作成
            calendar_master = None
            if calendar_id:
                if calendar_id not in calendar_master_map:
                    # ブランドコードを推測
                    brand = None
                    # 1001_SKAEC_A -> AECを抽出
                    parts = calendar_id.split('_')
                    if len(parts) >= 2:
                        possible_brand = parts[1].replace('SK', '')
                        brand = brand_map.get(possible_brand)

                    calendar_master, _ = CalendarMaster.objects.get_or_create(
                        tenant_id=tenant_id,
                        code=calendar_id,
                        defaults={
                            'name': calendar_id,
                            'brand': brand,
                            'lesson_type': CalendarMaster._detect_lesson_type(calendar_id),
                        }
                    )
                    calendar_master_map[calendar_id] = calendar_master
                else:
                    calendar_master = calendar_master_map[calendar_id]

            # 振替拒否判定
            is_makeup_allowed = rejection_code != 'C'

            # 既存チェック
            existing = LessonCalendar.objects.filter(
                tenant_id=tenant_id,
                calendar_code=calendar_id,
                lesson_date=lesson_date
            ).first()

            data = {
                'tenant_id': tenant_id,
                'calendar_master': calendar_master,
                'calendar_code': calendar_id,
                'lesson_date': lesson_date,
                'day_of_week': day_of_week,
                'is_open': is_open,
                'lesson_type': lesson_type,
                'display_label': display_label if display_label != '休' else '',
                'ticket_type': ticket_type,
                'ticket_sequence': int(ticket_sequence) if ticket_sequence else None,
                'is_makeup_allowed': is_makeup_allowed,
                'rejection_reason': rejection_reason,
                'ticket_issue_count': int(ticket_issue_count) if ticket_issue_count else None,
                'valid_days': int(valid_days) if valid_days else 90,
                'notice_message': notice,
                'auto_send_notice': auto_send.lower() in ['y', 'yes', '1', 'true'],
                'holiday_name': holiday_name,
            }

            if existing:
                for key, value in data.items():
                    setattr(existing, key, value)
                existing.save()
                updated += 1
            else:
                LessonCalendar.objects.create(**data)
                imported += 1

            if (imported + updated) % 100 == 0:
                print(f"  処理中... {imported + updated} 件")

        except Exception as e:
            errors.append(f"行 {row_num}: {calendar_id}/{lesson_date} - {str(e)}")
            continue

    return imported, updated, errors


def import_lesson_calendar(xlsx_path):
    """開講カレンダーをインポート"""

    # テナント取得
    tenant = Tenant.objects.filter(tenant_code='100000').first()
    if not tenant:
        print("エラー: アンイングリッシュグループテナントが見つかりません")
        sys.exit(1)
    tenant_id = tenant.id
    print(f"テナント: {tenant.tenant_name} ({tenant_id})")

    # ブランドマップ作成
    brand_map = {}
    for b in Brand.objects.filter(tenant_id=tenant_id):
        brand_map[b.brand_code] = b
    print(f"ブランドマップ: {len(brand_map)} 件")

    # カレンダーマスターマップ
    calendar_master_map = {}
    for cm in CalendarMaster.objects.filter(tenant_id=tenant_id):
        calendar_master_map[cm.code] = cm

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # インポート対象シート
    sheets = [
        '★新OZA授業カレンダー_もとA',
        '★新OZA授業カレンダー_もとB',
        '★新OZA授業カレンダー_もとP',
        '★新OZA授業カレンダー_Y',
    ]

    total_imported = 0
    total_updated = 0
    all_errors = []

    for sheet_name in sheets:
        if sheet_name in wb.sheetnames:
            print(f"\nシート処理中: {sheet_name}")
            ws = wb[sheet_name]
            imported, updated, errors = import_calendar_sheet(ws, tenant_id, brand_map, calendar_master_map)
            total_imported += imported
            total_updated += updated
            all_errors.extend(errors)
            print(f"  完了: 新規 {imported} 件, 更新 {updated} 件")

    return total_imported, total_updated, all_errors


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("使用方法: python scripts/import_lesson_calendar.py <xlsx_path>")
        sys.exit(1)

    xlsx_path = sys.argv[1]

    if not os.path.exists(xlsx_path):
        print(f"エラー: ファイルが見つかりません: {xlsx_path}")
        sys.exit(1)

    print(f"インポート開始: {xlsx_path}")
    print("-" * 50)

    imported, updated, errors = import_lesson_calendar(xlsx_path)

    print("-" * 50)
    print(f"完了!")
    print(f"  新規作成: {imported} 件")
    print(f"  更新: {updated} 件")

    if errors:
        print(f"  エラー: {len(errors)} 件")
        for error in errors[:10]:
            print(f"    - {error}")
        if len(errors) > 10:
            print(f"    ... 他 {len(errors) - 10} 件")
