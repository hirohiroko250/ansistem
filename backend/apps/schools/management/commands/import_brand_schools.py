"""
Import BrandSchool data from Excel file

Excelファイル（T14_開講時間割）からブランド-校舎の組み合わせを抽出し、
BrandSchoolテーブルにインポートする

Usage:
    python manage.py import_brand_schools --file path/to/excel.xlsx
    python manage.py import_brand_schools --file path/to/excel.xlsx --dry-run
    python manage.py import_brand_schools --file path/to/excel.xlsx --tenant default
"""
import re
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class Command(BaseCommand):
    help = 'ExcelファイルからBrandSchoolデータをインポート'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='/Users/hirosesuzu/Desktop/アンシステム/Claude-Code-Communication/instructions/OZA/T14_開講時間割_ 時間割_試作品.xlsx',
            help='Excelファイルのパス'
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='実際にはインポートせず、処理内容を表示のみ'
        )
        parser.add_argument(
            '--tenant',
            type=str,
            default='OZA',
            help='テナントコード（デフォルト: OZA=おざシステム）'
        )
        parser.add_argument(
            '--clear',
            action='store_true',
            help='既存のBrandSchoolデータをクリアしてから'
        )

    def handle(self, *args, **options):
        if load_workbook is None:
            raise CommandError('openpyxlがインストールされていません。pip install openpyxl を実行してください。')

        file_path = options['file']
        dry_run = options['dry_run']
        tenant_code = options['tenant']
        clear_existing = options['clear']

        self.stdout.write(f'Excelファイルを読み込み中: {file_path}')

        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            raise CommandError(f'Excelファイルの読み込みに失敗: {e}')

        # ブランドマッピングを作成（T14のB0001 -> DBのAEC）
        brand_mapping = self._build_brand_mapping(wb)
        self.stdout.write(f'ブランドマッピング: {len(brand_mapping)} 件')

        # T14からブランド-校舎組み合わせを抽出
        brand_school_pairs = self._extract_brand_school_pairs(wb)
        self.stdout.write(f'抽出したブランド-校舎組み合わせ: {len(brand_school_pairs)} 件')

        if dry_run:
            self.stdout.write(self.style.WARNING('=== ドライラン モード ==='))
            self._display_preview(brand_school_pairs, brand_mapping, tenant_code)
        else:
            self._import_data(brand_school_pairs, brand_mapping, tenant_code, clear_existing)

    def _build_brand_mapping(self, wb):
        """
        T12シートからブランドマッピングを作成
        T14のブランドID（B0001）-> DBのブランドコード（AEC）
        """
        mapping = {}
        sheet_name = '③ T12_ブランド情報　これを元へ張り付ける事'

        try:
            ws = wb[sheet_name]
        except KeyError:
            self.stdout.write(self.style.WARNING(f'シート "{sheet_name}" が見つかりません'))
            return mapping

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] and row[6]:  # 契約ブランドID と ClassブランドID
                contract_id = str(row[2]).strip()  # AEC, SOR等
                class_brand_id = str(row[6]).strip()  # B10001等

                # B10001 -> B0001 へ変換
                if class_brand_id.startswith('B1'):
                    t14_brand_id = 'B' + class_brand_id[2:].zfill(4)
                else:
                    t14_brand_id = class_brand_id

                # 最初のマッピング（主要ブランド）を優先し、上書きしない
                if t14_brand_id not in mapping:
                    mapping[t14_brand_id] = contract_id

        return mapping

    def _extract_brand_school_pairs(self, wb):
        """
        T14シートからブランド-校舎の組み合わせを抽出
        """
        pairs = set()
        sheet_name = 'T14_開講時間割_ 時間割Group版'

        try:
            ws = wb[sheet_name]
        except KeyError:
            self.stdout.write(self.style.WARNING(f'シート "{sheet_name}" が見つかりません'))
            return pairs

        for row in ws.iter_rows(min_row=2, values_only=True):
            schedule_id = str(row[0]) if row[0] else ''

            # B0001_T0002_C4145_S0001_10_4 のようなフォーマットをパース
            brand_match = re.search(r'B(\d+)', schedule_id)
            school_match = re.search(r'S(\d+)', schedule_id)

            if brand_match and school_match:
                brand_id = f'B{brand_match.group(1).zfill(4)}'
                # S0001 -> S10001 へ変換（DBの校舎コード形式）
                school_num = school_match.group(1)
                school_code = f'S1{school_num.zfill(4)}'
                pairs.add((brand_id, school_code))

        return pairs

    def _display_preview(self, pairs, brand_mapping, tenant_code):
        """ドライランでの内容表示"""
        from apps.schools.models import Brand, School
        from apps.tenants.models import Tenant

        try:
            tenant = Tenant.objects.get(tenant_code=tenant_code)
        except Tenant.DoesNotExist:
            self.stdout.write(self.style.ERROR(f'テナント "{tenant_code}" が見つかりません'))
            return

        # 既存のBrand/Schoolを取得
        brands = {b.brand_code: b for b in Brand.objects.filter(tenant_id=tenant.id)}
        schools = {s.school_code: s for s in School.objects.filter(tenant_id=tenant.id)}

        self.stdout.write('\n=== インポート予定データ ===')

        imported_count = 0
        skipped_brands = set()
        skipped_schools = set()

        for brand_id, school_code in sorted(pairs):
            brand_code = brand_mapping.get(brand_id)

            if not brand_code:
                skipped_brands.add(brand_id)
                continue

            if brand_code not in brands:
                skipped_brands.add(f'{brand_id}({brand_code})')
                continue

            if school_code not in schools:
                skipped_schools.add(school_code)
                continue

            brand = brands[brand_code]
            school = schools[school_code]
            self.stdout.write(f'  {brand.brand_name} - {school.school_name}')
            imported_count += 1

        self.stdout.write(f'\n合計: {imported_count} 件がインポートされます')

        if skipped_brands:
            self.stdout.write(self.style.WARNING(f'スキップされたブランド: {sorted(skipped_brands)}'))
        if skipped_schools:
            self.stdout.write(self.style.WARNING(f'スキップされた校舎: {sorted(skipped_schools)}'))

    @transaction.atomic
    def _import_data(self, pairs, brand_mapping, tenant_code, clear_existing):
        """データをインポート"""
        from apps.schools.models import Brand, School, BrandSchool
        from apps.tenants.models import Tenant

        try:
            tenant = Tenant.objects.get(tenant_code=tenant_code)
        except Tenant.DoesNotExist:
            raise CommandError(f'テナント "{tenant_code}" が見つかりません')

        # 既存のBrand/Schoolを取得
        brands = {b.brand_code: b for b in Brand.objects.filter(tenant_id=tenant.id)}
        schools = {s.school_code: s for s in School.objects.filter(tenant_id=tenant.id)}

        if clear_existing:
            deleted_count = BrandSchool.objects.filter(tenant_id=tenant.id).delete()[0]
            self.stdout.write(f'既存データ {deleted_count} 件を削除しました')

        created_count = 0
        skipped_count = 0
        skipped_brands = set()
        skipped_schools = set()

        for brand_id, school_code in pairs:
            brand_code = brand_mapping.get(brand_id)

            if not brand_code:
                skipped_brands.add(brand_id)
                skipped_count += 1
                continue

            if brand_code not in brands:
                skipped_brands.add(f'{brand_id}({brand_code})')
                skipped_count += 1
                continue

            if school_code not in schools:
                skipped_schools.add(school_code)
                skipped_count += 1
                continue

            brand = brands[brand_code]
            school = schools[school_code]

            # 既に存在するかチェック
            obj, created = BrandSchool.objects.get_or_create(
                tenant_id=tenant.id,
                brand=brand,
                school=school,
                defaults={
                    'is_active': True,
                    'sort_order': 0,
                }
            )

            if created:
                created_count += 1
                self.stdout.write(f'  作成: {brand.brand_name} - {school.school_name}')

        self.stdout.write(self.style.SUCCESS(f'\n完了: {created_count} 件作成, {skipped_count} 件スキップ'))

        if skipped_brands:
            self.stdout.write(self.style.WARNING(f'スキップされたブランド: {sorted(skipped_brands)}'))
        if skipped_schools:
            self.stdout.write(self.style.WARNING(f'スキップされた校舎: {sorted(skipped_schools)}'))
